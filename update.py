#!/usr/bin/env python3
"""
PCG Shipping Rule - Data Update Script
=======================================
Usage:
    python update.py                        # uses "PCG SP rule.xlsx" in the same directory
    python update.py "other_file.xlsx"      # specify a different file

The script reads the Excel file, extracts data from the "PCG" sheet,
and regenerates index.html with the latest data.

Requirements:
    pip install openpyxl
"""

import sys
import json
import os
import subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "template.html")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "index.html")
DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "PCG SP rule.xlsx")

# Columns to extract (index in Excel row -> clean key name)
HEADER_MAP = {
    2:  "ship to code",
    3:  "ship to name",
    4:  "country",
    5:  "sub GEO",
    6:  "GEO",
    7:  "Term",
    8:  "Incoterm",
    9:  "Logistics Focal",
    10: "Hic Focal",
    11: "货代/规则指定人 (包括FF/Planning/运输team）",
    12: "普货合并规则",
    13: "卡控不出 （仓库）",
    14: "GL（客服）",
    15: "CC GL",
    16: "快递账号",
    17: "货代账号",
    18: "SinglePack",
    19: "Brand Model",
    20: "Palletization",
    21: "PP Flag YES 快递( DG不适用)",
    22: "PP Flag YES 空运货代 (部分DG不适用）",
    23: "普货-20",
    24: "普货-40",
    25: "普货-60",
    26: "DG-20",
    27: "DG-40",
    28: "DG-60",
    29: "地址",
    30: "Booking/Greenlight 收件人",
    31: "备注",
}


def extract_data(excel_path):
    """Read the PCG sheet and return structured data."""
    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if "PCG" not in wb.sheetnames:
        print(f"Error: Sheet 'PCG' not found. Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["PCG"]
    data = []
    countries = set()
    geos = set()
    subgeos = set()
    terms = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip rows without a ship to code
        if not row[2]:
            continue

        record = {}
        for col_idx, key in HEADER_MAP.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                val = str(val).strip()
                if val.startswith("="):
                    val = ""
            else:
                val = ""
            record[key] = val

        if not record["ship to code"]:
            continue

        data.append(record)

        # Collect filter values
        if record["country"]:
            countries.add(record["country"])
        if record["GEO"]:
            geos.add(record["GEO"])
        if record["sub GEO"]:
            subgeos.add(record["sub GEO"])
        if record["Term"]:
            terms.add(record["Term"])

    wb.close()

    result = {
        "data": data,
        "filters": {
            "countries": sorted(countries),
            "geos": sorted(geos),
            "subgeos": sorted(subgeos),
            "terms": sorted(terms),
        },
    }
    print(f"Extracted {len(data)} records")
    print(f"  Countries: {len(countries)}, GEOs: {len(geos)}, SubGEOs: {len(subgeos)}, Terms: {len(terms)}")
    return result


def generate_html(data, update_date):
    """Inject data into the HTML template and write index.html."""
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: Template not found at {TEMPLATE_PATH}")
        sys.exit(1)

    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template = f.read()

    js_data = json.dumps(data, ensure_ascii=False)
    html = template.replace("__DATA_PLACEHOLDER__", js_data)

    # Update the date in the header badge
    html = html.replace("Updated __DATE_PLACEHOLDER__", f"Updated {update_date}")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    size_mb = os.path.getsize(OUTPUT_PATH) / (1024 * 1024)
    print(f"Generated: {OUTPUT_PATH} ({size_mb:.1f} MB)")


def push_to_github(update_date):
    """Commit and push the generated page when its content changed."""
    def run_git(*args, capture_output=False):
        return subprocess.run(
            ["git", *args],
            cwd=SCRIPT_DIR,
            check=True,
            text=True,
            capture_output=capture_output,
        )

    try:
        run_git("rev-parse", "--is-inside-work-tree", capture_output=True)
        branch = run_git("branch", "--show-current", capture_output=True).stdout.strip()
        if not branch:
            raise RuntimeError("Git is in detached HEAD state; cannot determine a branch to push.")

        run_git("add", "--", os.path.basename(OUTPUT_PATH))
        changed = subprocess.run(
            ["git", "diff", "--cached", "--quiet", "--", os.path.basename(OUTPUT_PATH)],
            cwd=SCRIPT_DIR,
        ).returncode != 0

        if not changed:
            print("No changes to index.html; GitHub push skipped.")
            return

        run_git("commit", "-m", f"Update shipping rules for {update_date}")
        run_git("push", "origin", branch)
        print(f"Pushed index.html to GitHub branch '{branch}'.")
    except (OSError, RuntimeError, subprocess.CalledProcessError) as error:
        print(f"Error: GitHub push failed: {error}")
        sys.exit(1)


def main():
    excel_path = sys.argv[1] if len(sys.argv) >= 2 else DEFAULT_EXCEL

    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        print(f"Place your Excel file as: {DEFAULT_EXCEL}")
        sys.exit(1)

    update_date = datetime.now().strftime("%Y.%m.%d")

    data = extract_data(excel_path)
    generate_html(data, update_date)
    push_to_github(update_date)
    print("Done!")


if __name__ == "__main__":
    main()
